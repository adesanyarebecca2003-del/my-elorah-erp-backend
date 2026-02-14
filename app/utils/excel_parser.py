from openpyxl import load_workbook


def parse_excel(file):
    wb = load_workbook(file.file, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    return rows